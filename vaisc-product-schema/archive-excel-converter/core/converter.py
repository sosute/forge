"""
Excel to Markdown converter using MarkItDown
"""

import os
import subprocess
import tempfile
from pathlib import Path
from typing import List, Dict, Optional, Tuple
from dataclasses import dataclass
import pandas as pd
import openpyxl

from .config import Config


@dataclass
class SheetInfo:
    """Excel sheet information"""
    name: str
    index: int
    rows: int
    cols: int
    non_empty_cells: int
    data_quality_score: float = 0.0
    issues: List[str] = None
    
    def __post_init__(self):
        if self.issues is None:
            self.issues = []


@dataclass
class ConversionResult:
    """Conversion operation result"""
    success: bool
    markdown_content: str = ""
    sheets: List[SheetInfo] = None
    error_message: str = ""
    file_path: str = ""
    
    def __post_init__(self):
        if self.sheets is None:
            self.sheets = []


class ExcelConverter:
    """Excel to Markdown converter using MarkItDown"""
    
    def __init__(self, config: Optional[Config] = None):
        self.config = config or Config()
        self._check_markitdown()
    
    def _check_markitdown(self) -> None:
        """Check if MarkItDown is available"""
        try:
            result = subprocess.run(
                ['python', '-m', 'markitdown', '--version'],
                capture_output=True,
                text=True,
                timeout=10
            )
            if result.returncode != 0:
                raise RuntimeError("MarkItDown is not properly installed")
        except (subprocess.TimeoutExpired, FileNotFoundError) as e:
            raise RuntimeError(f"MarkItDown is not available: {e}")
    
    def convert_excel_to_markdown(self, excel_path: str, output_path: Optional[str] = None, 
                                 target_sheets: Optional[List[str]] = None) -> ConversionResult:
        """
        Convert Excel file to Markdown using MarkItDown
        
        Args:
            excel_path: Path to Excel file
            output_path: Optional output file path
            target_sheets: Optional list of sheet names to convert (None = all sheets)
            
        Returns:
            ConversionResult with conversion status and content
        """
        try:
            # Validate input file
            if not os.path.exists(excel_path):
                return ConversionResult(
                    success=False,
                    error_message=f"Excel file not found: {excel_path}"
                )
            
            # Get sheet information before conversion
            sheets = self.get_sheet_info(excel_path)
            
            # Handle specific sheet selection
            if target_sheets:
                # Validate target sheets exist
                available_sheets = [sheet.name for sheet in sheets]
                invalid_sheets = [name for name in target_sheets if name not in available_sheets]
                if invalid_sheets:
                    return ConversionResult(
                        success=False,
                        error_message=f"Sheet(s) not found: {', '.join(invalid_sheets)}. Available: {', '.join(available_sheets)}"
                    )
                
                # Convert specific sheets only
                return self._convert_specific_sheets(excel_path, target_sheets, sheets, output_path)
            
            # Convert using MarkItDown (all sheets)
            if output_path:
                # Convert to file
                result = subprocess.run(
                    ['python', '-m', 'markitdown', excel_path, '-o', output_path],
                    capture_output=True,
                    text=True,
                    timeout=self.config.processing.timeout_seconds
                )
                
                if result.returncode == 0:
                    with open(output_path, 'r', encoding='utf-8') as f:
                        content = f.read()
                else:
                    return ConversionResult(
                        success=False,
                        error_message=f"MarkItDown conversion failed: {result.stderr}"
                    )
            else:
                # Convert to stdout
                result = subprocess.run(
                    ['python', '-m', 'markitdown', excel_path],
                    capture_output=True,
                    text=True,
                    timeout=self.config.processing.timeout_seconds
                )
                
                if result.returncode == 0:
                    content = result.stdout
                else:
                    return ConversionResult(
                        success=False,
                        error_message=f"MarkItDown conversion failed: {result.stderr}"
                    )
            
            return ConversionResult(
                success=True,
                markdown_content=content,
                sheets=sheets,
                file_path=output_path or ""
            )
            
        except subprocess.TimeoutExpired:
            return ConversionResult(
                success=False,
                error_message="Conversion timeout exceeded"
            )
        except Exception as e:
            return ConversionResult(
                success=False,
                error_message=f"Conversion error: {str(e)}"
            )
    
    def get_sheet_info(self, excel_path: str) -> List[SheetInfo]:
        """
        Analyze Excel file and get sheet information
        
        Args:
            excel_path: Path to Excel file
            
        Returns:
            List of SheetInfo objects
        """
        sheets = []
        
        try:
            # Use openpyxl for .xlsx files
            if excel_path.endswith('.xlsx'):
                wb = openpyxl.load_workbook(excel_path, data_only=True)
                for idx, sheet_name in enumerate(wb.sheetnames):
                    ws = wb[sheet_name]
                    
                    # Count non-empty cells
                    non_empty = 0
                    total_cells = 0
                    for row in ws.iter_rows():
                        for cell in row:
                            total_cells += 1
                            if cell.value is not None and str(cell.value).strip():
                                non_empty += 1
                    
                    # Calculate data quality score
                    quality_score = non_empty / max(total_cells, 1) if total_cells > 0 else 0
                    
                    # Detect issues
                    issues = self._detect_sheet_issues(ws)
                    
                    sheet_info = SheetInfo(
                        name=sheet_name,
                        index=idx,
                        rows=ws.max_row,
                        cols=ws.max_column,
                        non_empty_cells=non_empty,
                        data_quality_score=quality_score,
                        issues=issues
                    )
                    sheets.append(sheet_info)
                
                wb.close()
            
            # Use pandas for .xls files
            elif excel_path.endswith('.xls'):
                xls = pd.ExcelFile(excel_path)
                for idx, sheet_name in enumerate(xls.sheet_names):
                    df = pd.read_excel(xls, sheet_name=sheet_name)
                    
                    non_empty = df.count().sum()
                    total_cells = df.size
                    quality_score = non_empty / max(total_cells, 1)
                    
                    issues = self._detect_dataframe_issues(df)
                    
                    sheet_info = SheetInfo(
                        name=sheet_name,
                        index=idx,
                        rows=len(df),
                        cols=len(df.columns),
                        non_empty_cells=int(non_empty),
                        data_quality_score=quality_score,
                        issues=issues
                    )
                    sheets.append(sheet_info)
        
        except Exception as e:
            # Return basic info if detailed analysis fails
            sheets.append(SheetInfo(
                name="Unknown",
                index=0,
                rows=0,
                cols=0,
                non_empty_cells=0,
                issues=[f"Analysis error: {str(e)}"]
            ))
        
        return sheets
    
    def _detect_sheet_issues(self, worksheet) -> List[str]:
        """Detect common issues in Excel worksheet"""
        issues = []
        
        # Check for merged cells
        if worksheet.merged_cells.ranges:
            issues.append(f"Contains {len(worksheet.merged_cells.ranges)} merged cell ranges")
        
        # Check for formulas
        formula_count = 0
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.data_type == 'f':  # Formula
                    formula_count += 1
        
        if formula_count > 0:
            issues.append(f"Contains {formula_count} formula cells")
        
        # Check for very wide sheets (potential formatting issues)
        if worksheet.max_column > 50:
            issues.append(f"Very wide sheet ({worksheet.max_column} columns)")
        
        return issues
    
    def _detect_dataframe_issues(self, df: pd.DataFrame) -> List[str]:
        """Detect common issues in pandas DataFrame"""
        issues = []
        
        # Check for unnamed columns
        unnamed_cols = [col for col in df.columns if 'Unnamed:' in str(col)]
        if unnamed_cols:
            issues.append(f"Contains {len(unnamed_cols)} unnamed columns")
        
        # Check for high NaN ratio
        nan_ratio = df.isnull().sum().sum() / df.size
        if nan_ratio > 0.5:
            issues.append(f"High NaN ratio ({nan_ratio:.1%})")
        
        # Check for very sparse data
        if df.count().sum() / df.size < 0.1:
            issues.append("Very sparse data (>90% empty)")
        
        return issues
    
    def validate_excel_file(self, excel_path: str) -> Tuple[bool, str]:
        """
        Validate Excel file before conversion
        
        Args:
            excel_path: Path to Excel file
            
        Returns:
            Tuple of (is_valid, error_message)
        """
        if not os.path.exists(excel_path):
            return False, f"File not found: {excel_path}"
        
        if not excel_path.lower().endswith(('.xlsx', '.xls')):
            return False, "File must be .xlsx or .xls format"
        
        # Check file size
        file_size = os.path.getsize(excel_path)
        max_size = self.config.processing.memory_limit_mb * 1024 * 1024
        if file_size > max_size:
            return False, f"File too large: {file_size / 1024 / 1024:.1f}MB (max: {self.config.processing.memory_limit_mb}MB)"
        
        # Try to open file
        try:
            if excel_path.endswith('.xlsx'):
                wb = openpyxl.load_workbook(excel_path, read_only=True)
                wb.close()
            else:
                pd.ExcelFile(excel_path).close()
        except Exception as e:
            return False, f"Cannot open Excel file: {str(e)}"
        
        return True, "File is valid"
    
    def _convert_specific_sheets(self, excel_path: str, target_sheets: List[str], 
                               all_sheets: List[SheetInfo], output_path: Optional[str] = None) -> ConversionResult:
        """
        Convert specific sheets from Excel file to Markdown
        
        Args:
            excel_path: Path to Excel file
            target_sheets: List of sheet names to convert
            all_sheets: All sheet information
            output_path: Optional output file path
            
        Returns:
            ConversionResult with conversion status and content
        """
        try:
            # Filter sheets to target ones only
            target_sheet_info = [sheet for sheet in all_sheets if sheet.name in target_sheets]
            
            # Create temporary Excel file with only target sheets
            temp_excel_path = self._create_filtered_excel(excel_path, target_sheets)
            
            try:
                # Convert filtered Excel using MarkItDown
                if output_path:
                    result = subprocess.run(
                        ['python', '-m', 'markitdown', temp_excel_path, '-o', output_path],
                        capture_output=True,
                        text=True,
                        timeout=self.config.processing.timeout_seconds
                    )
                    
                    if result.returncode == 0:
                        with open(output_path, 'r', encoding='utf-8') as f:
                            content = f.read()
                    else:
                        return ConversionResult(
                            success=False,
                            error_message=f"MarkItDown conversion failed: {result.stderr}"
                        )
                else:
                    result = subprocess.run(
                        ['python', '-m', 'markitdown', temp_excel_path],
                        capture_output=True,
                        text=True,
                        timeout=self.config.processing.timeout_seconds
                    )
                    
                    if result.returncode == 0:
                        content = result.stdout
                    else:
                        return ConversionResult(
                            success=False,
                            error_message=f"MarkItDown conversion failed: {result.stderr}"
                        )
                
                return ConversionResult(
                    success=True,
                    markdown_content=content,
                    sheets=target_sheet_info,
                    file_path=output_path or ""
                )
                
            finally:
                # Clean up temporary file
                if os.path.exists(temp_excel_path):
                    os.remove(temp_excel_path)
                    
        except Exception as e:
            return ConversionResult(
                success=False,
                error_message=f"Sheet filtering error: {str(e)}"
            )
    
    def _create_filtered_excel(self, excel_path: str, target_sheets: List[str]) -> str:
        """
        Create temporary Excel file containing only target sheets
        
        Args:
            excel_path: Original Excel file path
            target_sheets: List of sheet names to include
            
        Returns:
            Path to temporary Excel file
        """
        import tempfile
        import shutil
        
        # Create temporary file
        temp_dir = tempfile.mkdtemp()
        temp_excel_path = os.path.join(temp_dir, f"filtered_{os.path.basename(excel_path)}")
        
        if excel_path.endswith('.xlsx'):
            # Use openpyxl for .xlsx files
            import openpyxl
            
            # Load original workbook
            wb_orig = openpyxl.load_workbook(excel_path)
            
            # Create new workbook
            wb_new = openpyxl.Workbook()
            
            # Remove default sheet
            wb_new.remove(wb_new.active)
            
            # Copy target sheets
            for sheet_name in target_sheets:
                if sheet_name in wb_orig.sheetnames:
                    # Copy sheet
                    ws_orig = wb_orig[sheet_name]
                    ws_new = wb_new.create_sheet(title=sheet_name)
                    
                    # Copy data
                    for row in ws_orig.iter_rows():
                        for cell in row:
                            ws_new[cell.coordinate].value = cell.value
            
            # Save new workbook
            wb_new.save(temp_excel_path)
            wb_orig.close()
            wb_new.close()
            
        else:
            # Use pandas for .xls files
            # Read all sheets
            all_sheets = pd.read_excel(excel_path, sheet_name=None)
            
            # Filter to target sheets
            filtered_sheets = {name: df for name, df in all_sheets.items() if name in target_sheets}
            
            # Write to new file
            with pd.ExcelWriter(temp_excel_path, engine='openpyxl') as writer:
                for sheet_name, df in filtered_sheets.items():
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        return temp_excel_path
    
    def list_sheet_names(self, excel_path: str) -> List[str]:
        """
        Get list of sheet names from Excel file
        
        Args:
            excel_path: Path to Excel file
            
        Returns:
            List of sheet names
        """
        try:
            if excel_path.endswith('.xlsx'):
                wb = openpyxl.load_workbook(excel_path, read_only=True)
                sheet_names = wb.sheetnames
                wb.close()
            else:
                xls = pd.ExcelFile(excel_path)
                sheet_names = xls.sheet_names
                xls.close()
            
            return sheet_names
            
        except Exception as e:
            print(f"Warning: Could not read sheet names: {e}")
            return []